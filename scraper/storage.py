"""
Data persistence layer.

Covers:
  - CSV output (default, zero extra dependencies)
  - Excel output with styled headers (requires openpyxl)
  - Checkpoint read/write (atomic, crash-safe via temp-file + os.replace)
  - Done-queries log (lightweight resume without a full checkpoint)
  - Deduplication set reconstruction from existing output files
"""

from __future__ import annotations

import csv
import json
import logging
import os
import re
from datetime import date
from pathlib import Path
from typing import Any

log = logging.getLogger("maps_scraper")

# ── Output schema ──────────────────────────────────────────────────────────────
# Single source of truth for column names — shared by all output and filter code.

OUTPUT_FIELDS: list[str] = [
    "Company Name",
    "Phone",
    "Website",
    "Email",
    "Postcode",
    "Category",
    "Rating",
    "Address",
    "Email Status",
    "Phone Status",
    "Source",
]

_HEADER_BG = "2E4057"   # dark navy header fill (Excel only)
_HEADER_FG = "FFFFFF"
_COL_WIDTH  = 24

# Checkpoint format version — increment when the schema changes so old
# checkpoints are detected and gracefully rejected rather than silently
# producing corrupt resumes.
CHECKPOINT_VERSION = 2


# ── Output path ───────────────────────────────────────────────────────────────

def build_output_path(config: dict[str, Any]) -> Path:
    """
    Build a timestamped output file path from the active config.

    The filename encodes the search query + location + date so that
    separate runs never overwrite each other.

    Args:
        config: Full config dict.

    Returns:
        Path object pointing to the output file (not yet created).
    """
    out = config.get("output", {})
    d   = Path(out.get("directory", "output"))
    d.mkdir(parents=True, exist_ok=True)
    pfx = out.get("filename_prefix", "MapsScrape")
    q   = re.sub(r"[^\w\s]", "", config["search"]["query"]).strip().replace(" ", "_")
    q   = re.sub(r"_+", "_", q)
    loc = re.sub(r"\s+", "_", config["search"]["location"])
    fmt = out.get("format", "csv")
    ext = "xlsx" if fmt == "excel" else "csv"
    return d / f"{pfx}_{q}_{loc}_{date.today():%Y%m%d}.{ext}"


# ── Row builder ───────────────────────────────────────────────────────────────

def build_row(
    place: dict,
    email: str,
    phone: str,
    category: str,
    config: dict[str, Any],
) -> dict:
    """
    Assemble a single output row dict from a scraped place and enrichment data.

    Applies CSV-injection sanitization to free-text fields (name, address)
    so that opening the output in Excel cannot trigger formula execution.

    Args:
        place:    Raw place dict from extract_place().
        email:    Email address found during enrichment (may be empty).
        phone:    Best phone number found (Maps listing or business website).
        category: Keyword-classified label from classify_company().
        config:   Full config dict (used for postcode extraction).

    Returns:
        Dict matching OUTPUT_FIELDS column order.
    """
    from .filters import extract_postcode, sanitize_cell

    name    = sanitize_cell(place.get("name", ""))
    address = sanitize_cell(place.get("address", ""))

    return {
        "Company Name": name,
        "Phone":        phone,
        "Website":      place.get("website", ""),
        "Email":        email,
        "Postcode":     extract_postcode(address, config),
        "Category":     category,
        "Rating":       place.get("rating", ""),
        "Address":      address,
        "Email Status": "found" if email else "notfound",
        "Phone Status": "found" if phone  else "notfound",
        "Source":       "Google Maps",
    }


# ── Save / load ───────────────────────────────────────────────────────────────

def load_existing_output(path: Path) -> list[dict]:
    """
    Load rows from a previous run's output file if it exists.

    Args:
        path: Output file path.

    Returns:
        List of row dicts, or empty list if file absent or unreadable.
    """
    if not path.exists():
        return []
    try:
        if path.suffix == ".xlsx":
            return _load_xlsx(path)
        return _load_csv(path)
    except Exception as exc:
        log.warning("Could not load existing output (%s): %s", path.name, exc)
        return []


def save_output(rows: list[dict], path: Path, fmt: str = "csv") -> None:
    """
    Write all result rows to a CSV or Excel file.

    Falls back to CSV automatically if openpyxl is not installed and logs
    a clear install hint rather than raising a bare ImportError.

    Args:
        rows: List of row dicts to write.
        path: Destination file path.
        fmt:  "csv" or "excel".
    """
    if not rows:
        return
    if fmt == "excel":
        try:
            import openpyxl  # noqa: F401  — check before attempting write
        except ImportError:
            log.warning(
                "⚠️  openpyxl is not installed — falling back to CSV.\n"
                "    To enable Excel output run:  pip install openpyxl"
            )
            path = path.with_suffix(".csv")
            _save_csv(rows, path)
            return
        try:
            _save_xlsx(rows, path)
            return
        except Exception as exc:
            log.warning("Excel write failed (%s) — falling back to CSV", exc)
            path = path.with_suffix(".csv")
    _save_csv(rows, path)


# ── CSV helpers ───────────────────────────────────────────────────────────────

def _save_csv(rows: list[dict], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def append_rows(new_rows: list[dict], path: Path) -> None:
    """
    Append only the new rows — O(new_rows) not O(total_rows).

    Uses append mode to avoid rewriting the entire file on every zone.
    Writes the CSV header only when the file does not exist or is empty.

    Args:
        new_rows: List of new row dicts to append.
        path:     Destination CSV file path.
    """
    if not new_rows:
        return
    write_header = not path.exists() or path.stat().st_size == 0
    with open(path, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS, extrasaction="ignore")
        if write_header:
            writer.writeheader()
        writer.writerows(new_rows)


def _load_csv(path: Path) -> list[dict]:
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _save_xlsx(rows: list[dict], path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(bold=True, color=_HEADER_FG)
    center      = Alignment(horizontal="center", vertical="center")

    ws.append(OUTPUT_FIELDS)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = _COL_WIDTH

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(f, "") for f in OUTPUT_FIELDS])

    wb.save(path)


def _load_xlsx(path: Path) -> list[dict]:
    import openpyxl
    wb      = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws      = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows    = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v for v in row):
            rows.append({h: (v or "") for h, v in zip(headers, row)})
    wb.close()
    return rows


# ── Checkpoint (atomic write) ─────────────────────────────────────────────────

def save_checkpoint(data: dict, config: dict[str, Any]) -> None:
    """
    Atomically write checkpoint data to disk.

    Uses a temp-file + os.replace pattern so a crash mid-write never
    produces a corrupt or empty checkpoint file. Injects CHECKPOINT_VERSION
    into every saved dict so future format changes can be detected on load.

    Args:
        data:   Serialisable dict (e.g. {"jobs": [...], "output_path": "..."}).
        config: Full config dict (reads files.checkpoint path).
    """
    path = config["files"]["checkpoint"]
    tmp  = path + ".tmp"
    payload = {"version": CHECKPOINT_VERSION, **data}
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f)
        os.replace(tmp, path)
    except OSError as exc:
        log.warning("Checkpoint save failed: %s", exc)


def load_checkpoint(config: dict[str, Any]) -> dict | None:
    """
    Load and return the checkpoint dict, or None if absent, corrupt, or
    from an incompatible version.

    Args:
        config: Full config dict.
    """
    path = config["files"]["checkpoint"]
    if not os.path.exists(path):
        return None
    try:
        with open(path, encoding="utf-8") as _f:
            content = _f.read().strip()
        if not content:
            return None
        data = json.loads(content)
        if data.get("version", 1) != CHECKPOINT_VERSION:
            log.warning(
                "Checkpoint version mismatch (got v%s, expected v%s) — "
                "ignoring stale checkpoint and starting fresh.",
                data.get("version", "?"), CHECKPOINT_VERSION,
            )
            return None
        return data
    except Exception as exc:
        log.warning("Checkpoint unreadable (%s) — starting fresh", exc)
        return None


def clear_checkpoint(config: dict[str, Any]) -> None:
    """Remove the checkpoint file if it exists."""
    path = config["files"]["checkpoint"]
    if os.path.exists(path):
        os.remove(path)


# ── Done-queries log ──────────────────────────────────────────────────────────

def log_done_query(query: str, config: dict[str, Any]) -> None:
    """Append a completed query string to the done-queries log."""
    path = config["files"]["done_queries"]
    with open(path, "a", encoding="utf-8") as f:
        f.write(query + "\n")


def load_done_queries(config: dict[str, Any]) -> set[str]:
    """Return the set of previously completed query strings."""
    path = config["files"]["done_queries"]
    if not os.path.exists(path):
        return set()
    with open(path, encoding="utf-8") as f:
        return {line.strip() for line in f if line.strip()}


def clear_done_queries(config: dict[str, Any]) -> None:
    """Remove the done-queries log if it exists."""
    path = config["files"]["done_queries"]
    if os.path.exists(path):
        os.remove(path)
